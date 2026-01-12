from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime
import models, auth, database
import io
import openpyxl
from openpyxl.styles import Font

router = APIRouter(prefix="/data", tags=["Data Management"])

# --- EXPORT TO EXCEL ---
@router.get("/export_excel")
def export_excel(db: Session = Depends(database.get_db), current_user: models.User = Depends(auth.get_current_user)):
    wb = openpyxl.Workbook()
    
    # 1. Transactions Sheet
    ws_txns = wb.active
    ws_txns.title = "Transactions"
    ws_txns.append(["Date", "Card Name", "Type", "Mode", "Amount", "Description", "Category", "EMI?", "Tenure"])
    
    txns = db.query(models.Transaction).join(models.Card).filter(models.Card.owner_id == current_user.id).all()
    for t in txns:
        ws_txns.append([
            t.date.strftime("%Y-%m-%d"),
            t.card.name,
            t.type,
            t.mode,
            t.amount,
            t.description,
            t.tag.name if t.tag else "",
            "Yes" if t.is_emi else "No",
            t.emi_tenure if t.is_emi else ""
        ])

    # 2. Cards Sheet
    ws_cards = wb.create_sheet("Cards")
    ws_cards.append(["Name", "Bank", "Network", "Last 4", "Total Limit", "Manual Limit", "Statement Day", "Due Day", "Card Type", "Expiry"])
    
    for c in current_user.cards:
        ws_cards.append([
            c.name, c.bank, c.network, c.last_4, c.total_limit, c.manual_limit, 
            c.statement_date, c.payment_due_date, c.card_type, c.expiry_date
        ])

    # 3. Lending Sheet
    ws_lend = wb.create_sheet("Lending")
    ws_lend.append(["Borrower", "Amount", "Lent Date", "Returned?", "Returned Date"])
    
    for l in current_user.lending:
        ws_lend.append([
            l.borrower_name, l.amount, l.lent_date.strftime("%Y-%m-%d"),
            "Yes" if l.is_returned else "No",
            l.returned_date.strftime("%Y-%m-%d") if l.returned_date else ""
        ])

    # 4. Income Sheet (Salary)
    ws_inc = wb.create_sheet("Income")
    ws_inc.append(["Company", "Joining Date", "Salary Date", "Amount"])
    
    companies = db.query(models.Company).filter(models.Company.owner_id == current_user.id).all()
    for comp in companies:
        for sal in comp.salaries:
            ws_inc.append([
                comp.name, comp.joining_date.strftime("%Y-%m-%d"),
                sal.date.strftime("%Y-%m-%d"), sal.amount
            ])

    # Style Headers
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"cc_track_backup_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# --- IMPORT FROM EXCEL ---
@router.post("/import_excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(database.get_db), current_user: models.User = Depends(auth.get_current_user)):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        
        # 1. Import Cards First (Dependencies)
        if "Cards" in wb.sheetnames:
            ws = wb["Cards"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue # Skip empty
                # Check if card exists
                exists = db.query(models.Card).filter(models.Card.name == row[0], models.Card.owner_id == current_user.id).first()
                if not exists:
                    new_card = models.Card(
                        name=row[0], bank=row[1], network=row[2], last_4=str(row[3]) if row[3] else None,
                        total_limit=float(row[4] or 0), manual_limit=float(row[5] or 0),
                        statement_date=int(row[6] or 1), payment_due_date=int(row[7] or 10),
                        card_type=row[8] or "Credit Card", expiry_date=str(row[9]) if row[9] else None,
                        owner_id=current_user.id
                    )
                    db.add(new_card)
            db.commit()

        # 2. Import Transactions
        if "Transactions" in wb.sheetnames:
            ws = wb["Transactions"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                # Find Card ID
                card = db.query(models.Card).filter(models.Card.name == row[1], models.Card.owner_id == current_user.id).first()
                if card:
                    # Tag Logic
                    tag_id = None
                    if row[6]:
                        tag_name = str(row[6]).strip().title()
                        tag = db.query(models.Tag).filter(models.Tag.name == tag_name, models.Tag.owner_id == current_user.id).first()
                        if not tag:
                            tag = models.Tag(name=tag_name, owner_id=current_user.id)
                            db.add(tag)
                            db.commit()
                        tag_id = tag.id
                    
                    # Avoid Duplicates (Simple check: same date, amount, description)
                    txn_date = row[0] if isinstance(row[0], datetime) else datetime.strptime(str(row[0]), "%Y-%m-%d")
                    exists = db.query(models.Transaction).filter(
                        models.Transaction.card_id == card.id,
                        models.Transaction.date == txn_date,
                        models.Transaction.amount == float(row[4]),
                        models.Transaction.description == row[5]
                    ).first()
                    
                    if not exists:
                        new_txn = models.Transaction(
                            date=txn_date,
                            type=row[2],
                            mode=row[3],
                            amount=float(row[4]),
                            description=row[5],
                            tag_id=tag_id,
                            is_emi=(row[7] == "Yes"),
                            emi_tenure=int(row[8]) if row[8] else None,
                            card_id=card.id
                        )
                        db.add(new_txn)
            db.commit()
            
        # 3. Import Lending
        if "Lending" in wb.sheetnames:
            ws = wb["Lending"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                lent_date = row[2] if isinstance(row[2], datetime) else datetime.strptime(str(row[2]), "%Y-%m-%d")
                
                # Check duplicate
                exists = db.query(models.Lending).filter(
                    models.Lending.borrower_name == row[0],
                    models.Lending.amount == float(row[1]),
                    models.Lending.lent_date == lent_date,
                    models.Lending.owner_id == current_user.id
                ).first()
                
                if not exists:
                    ret_date = None
                    if row[4]:
                        ret_date = row[4] if isinstance(row[4], datetime) else datetime.strptime(str(row[4]), "%Y-%m-%d")

                    new_lend = models.Lending(
                        borrower_name=row[0],
                        amount=float(row[1]),
                        lent_date=lent_date,
                        is_returned=(row[3] == "Yes"),
                        returned_date=ret_date,
                        owner_id=current_user.id
                    )
                    db.add(new_lend)
            db.commit()

        return {"message": "Import successful"}

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")